import datetime
import os

from django.http import HttpResponse
from django.shortcuts import render
from django.template import loader
from django.views.decorators.csrf import csrf_protect
from openpyxl import load_workbook
from django.http import QueryDict


from .buscadorDIAN import buscadorDIAN
from .forms import ExcelForm


# Create your views here.


def vista(request):
    if request.method == 'POST':
        form = ExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            workbook = load_workbook(archivo_excel)
            # Obtiene la primera hoja del libro de trabajo
            sheet = workbook.active

            # Obtiene los valores de la primera columna en una lista
            #primera_columna = []
            #for row in sheet.iter_rows(values_only=True):
            #    primera_columna.append(row[0])

            array = []
            for row in sheet.iter_rows():  # Itera sobre las filas de la hoja de trabajo
                for cell in row:  # Itera sobre las celdas de cada fila
                    if cell.value is not None:  # Verifica si la celda no está vacía
                        array.append(cell.value)
            print(array)

            querydict = QueryDict(request.POST.urlencode())
            valor = querydict.get('csrfmiddlewaretoken')

            #print(primera_columna)
            #buscar_dian(primera_columna,valor)

            buscador = buscadorDIAN(array, valor)
            buscador.multiples_busquedas()
            nombre_retornado = buscador.archivo_excel()


            context = {
                'archivo': 'archivo',
                'exito': True,
                'nombre_archivo': nombre_retornado
            }

            return render(request, 'resultado.html', context)

    else:
        form = ExcelForm()
        return render(request, 'buscador.html', {'form': form})


def descargar_archivo(request):
    param = request.GET.get('parametro')
    #print(param)
    file_path = os.path.join(os.getcwd().split('buscador')[0] + '\\archivos\\'+param)

    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = 'attachment; filename="{0}"'.format(os.path.basename(file_path))
            return response
    else:
        return HttpResponse("El archivo no existe.")


def buscar_dian(data, nombre_archivo):
    buscador = buscadorDIAN(data, nombre_archivo)
    buscador.multiples_busquedas()
    buscador.archivo_excel()




